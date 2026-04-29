import tempfile
import unittest
import warnings
from pathlib import Path

pytest_import_error = None
try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    pytest_import_error = exc


@unittest.skipIf(pytest_import_error is not None, "openpyxl is not installed")
class WriterExcelTest(unittest.TestCase):
    def test_central_report_has_required_sheets(self):
        from oracle_pg_sync.reports.writer_excel import write_central_report_xlsx

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"

            write_central_report_xlsx(
                path,
                sync_rows=[{"table_name": "public.sample", "status": "SUCCESS", "rows_loaded": 10, "elapsed_seconds": 2}],
                checksum_rows=[{"table_name": "public.sample", "chunk_key": "table", "status": "MATCH"}],
                dependency_rows=[{"table_name": "public.sample", "object_type": "VIEW", "object_name": "sample_v"}],
                watermark_rows=[{"table_name": "public.sample", "value": "10"}],
                checkpoint_rows=[{"table_name": "public.sample", "chunk_key": "full", "status": "success"}],
                config_sanitized={"oracle": {"password": "****"}},
            )

            workbook = load_workbook(path)

        self.assertEqual(
            workbook.sheetnames,
            [
                "00_Dashboard",
                "01_Run_Summary",
                "02_Table_Sync_Status",
                "03_Rowcount_Compare",
                "04_Checksum_Result",
                "05_Column_Diff",
                "06_Index_Compare",
                "07_Object_Dependency",
                "08_LOB_Columns",
                "09_Failed_Tables",
                "10_Watermark",
                "11_Checkpoint",
                "12_Performance",
                "13_Errors",
                "14_Config",
            ],
        )
        self.assertEqual(workbook["00_Dashboard"].freeze_panes, "A2")
        self.assertIsNotNone(workbook["04_Checksum_Result"].auto_filter.ref)

    def test_long_cell_values_are_truncated_before_openpyxl_warning(self):
        from oracle_pg_sync.reports.writer_excel import EXCEL_CELL_MAX_CHARS, write_central_report_xlsx

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "report.xlsx"

            with warnings.catch_warnings(record=True) as caught:
                warnings.simplefilter("always")
                write_central_report_xlsx(
                    path,
                    sync_rows=[
                        {
                            "table_name": "public.sample",
                            "status": "FAILED",
                            "message": "x" * 95056,
                        }
                    ],
                )

            workbook = load_workbook(path)

        messages = [str(item.message) for item in caught]
        self.assertFalse(any("Cell contents too long" in message for message in messages))
        value = workbook["02_Table_Sync_Status"]["C2"].value
        self.assertLessEqual(len(value), EXCEL_CELL_MAX_CHARS)
        self.assertIn("truncated", value)


if __name__ == "__main__":
    unittest.main()
