from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill

EXCEL_CELL_MAX_CHARS = 32767
EXCEL_SAFE_CELL_CHARS = 32000


def write_inventory_xlsx(path: Path, inventory_rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    _dataframe(inventory_rows).to_excel(path, index=False, sheet_name="inventory")


def write_rows_xlsx(path: Path, rows: list[dict], *, sheet_name: str = "rows") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    _dataframe(rows).to_excel(path, index=False, sheet_name=sheet_name[:31])


def write_central_report_xlsx(
    path: Path,
    *,
    inventory_rows: list[dict] | None = None,
    column_diff_rows: list[dict] | None = None,
    type_mismatch_rows: list[dict] | None = None,
    sync_rows: list[dict] | None = None,
    checksum_rows: list[dict] | None = None,
    lob_rows: list[dict] | None = None,
    dependency_rows: list[dict] | None = None,
    dependency_summary_rows: list[dict] | None = None,
    maintenance_rows: list[dict] | None = None,
    watermark_rows: list[dict] | None = None,
    checkpoint_rows: list[dict] | None = None,
    config_sanitized: dict[str, Any] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    inventory_rows = inventory_rows or []
    column_diff_rows = column_diff_rows or []
    type_mismatch_rows = type_mismatch_rows or []
    sync_rows = sync_rows or []
    checksum_rows = checksum_rows or []
    lob_rows = lob_rows or [row for row in sync_rows if row.get("lob_columns_detected") or row.get("lob_type")]
    dependency_rows = dependency_rows or []
    dependency_summary_rows = dependency_summary_rows or []
    maintenance_rows = maintenance_rows or []
    watermark_rows = watermark_rows or []
    checkpoint_rows = checkpoint_rows or []
    table_status_rows = sync_rows or inventory_rows
    rowcount_rows = _rowcount_rows(sync_rows, inventory_rows)
    error_rows = _error_rows(sync_rows, maintenance_rows)
    sheets = {
        "00_Dashboard": [_dashboard_row(table_status_rows, checksum_rows, watermark_rows, checkpoint_rows, lob_rows)],
        "01_Run_Summary": [_run_summary_row(table_status_rows, checksum_rows, watermark_rows, checkpoint_rows)],
        "02_Table_Sync_Status": table_status_rows,
        "03_Rowcount_Compare": rowcount_rows,
        "04_Checksum_Result": checksum_rows,
        "05_Column_Diff": column_diff_rows + type_mismatch_rows,
        "06_Index_Compare": _filter_dependency_rows(dependency_rows, {"INDEX"}),
        "07_Object_Dependency": _object_dependency_sheet_rows(dependency_rows, dependency_summary_rows),
        "08_LOB_Columns": lob_rows,
        "09_Failed_Tables": [row for row in table_status_rows if str(row.get("status", "")).upper() in {"FAILED", "MISMATCH", "MISSING"}],
        "10_Watermark": watermark_rows,
        "11_Checkpoint": checkpoint_rows,
        "12_Performance": _performance_rows(sync_rows),
        "13_Errors": error_rows,
        "14_Config": _flatten_config(config_sanitized or {}),
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, rows in sheets.items():
            _dataframe(rows).to_excel(writer, index=False, sheet_name=name[:31])
        workbook = writer.book
        for worksheet in workbook.worksheets:
            _format_sheet(worksheet)


def _dashboard_row(
    table_rows: list[dict],
    checksum_rows: list[dict],
    watermark_rows: list[dict],
    checkpoint_rows: list[dict],
    lob_rows: list[dict] | None = None,
) -> dict[str, Any]:
    lob_rows = lob_rows or []
    return {
        "total_tables": len(table_rows),
        "success": sum(1 for row in table_rows if str(row.get("status", "")).upper() in {"SUCCESS", "MATCH"}),
        "failed": sum(1 for row in table_rows if str(row.get("status", "")).upper() in {"FAILED", "MISMATCH", "MISSING"}),
        "checksum_pass": sum(1 for row in checksum_rows if row.get("status") == "MATCH"),
        "checksum_fail": sum(1 for row in checksum_rows if row.get("status") == "MISMATCH"),
        "rows_processed": sum(int(row.get("rows_loaded") or 0) for row in table_rows),
        "watermark_updates": len(watermark_rows),
        "resume_usage": sum(1 for row in checkpoint_rows if row.get("chunk_key") or row.get("status")),
        "lob_heavy_tables": sum(1 for row in lob_rows if str(row.get("classification", "")).upper() == "LOB-HEAVY"),
        "slow_tables": sum(1 for row in table_rows if float(row.get("elapsed_seconds") or 0) >= 300),
    }


def _run_summary_row(
    table_rows: list[dict],
    checksum_rows: list[dict],
    watermark_rows: list[dict],
    checkpoint_rows: list[dict],
) -> dict[str, Any]:
    row = _dashboard_row(table_rows, checksum_rows, watermark_rows, checkpoint_rows)
    row["duration_seconds"] = round(sum(float(item.get("elapsed_seconds") or 0) for item in table_rows), 3)
    row["warning"] = sum(1 for item in table_rows if str(item.get("status", "")).upper() == "WARNING")
    row["dry_run"] = sum(1 for item in table_rows if str(item.get("status", "")).upper() == "DRY_RUN")
    return row


def _rowcount_rows(sync_rows: list[dict], inventory_rows: list[dict]) -> list[dict]:
    rows = sync_rows or inventory_rows
    return [
        {
            "table_name": row.get("table_name"),
            "oracle_row_count": row.get("oracle_row_count"),
            "postgres_row_count": row.get("postgres_row_count"),
            "row_count_match": row.get("row_count_match"),
            "status": row.get("status"),
        }
        for row in rows
    ]


def _filter_dependency_rows(rows: list[dict], object_types: set[str]) -> list[dict]:
    return [row for row in rows if str(row.get("object_type", "")).upper() in object_types]


def _object_dependency_sheet_rows(dependency_rows: list[dict], summary_rows: list[dict]) -> list[dict]:
    detail_rows = _filter_dependency_rows(
        dependency_rows,
        {"VIEW", "MATERIALIZED VIEW", "PROCEDURE", "FUNCTION", "PACKAGE", "PACKAGE BODY", "SEQUENCE"},
    )
    return [
        {"row_type": "summary", **row}
        for row in summary_rows
    ] + [{"row_type": "detail", **row} for row in detail_rows]


def _performance_rows(sync_rows: list[dict]) -> list[dict]:
    return [
        {
            "table_name": row.get("table_name"),
            "mode": row.get("mode"),
            "rows_loaded": row.get("rows_loaded"),
            "elapsed_seconds": row.get("elapsed_seconds"),
            "rows_per_second": _rows_per_second(row),
        }
        for row in sync_rows
    ]


def _rows_per_second(row: dict) -> float | None:
    elapsed = float(row.get("elapsed_seconds") or 0)
    if elapsed <= 0:
        return None
    return round(float(row.get("rows_loaded") or 0) / elapsed, 3)


def _error_rows(sync_rows: list[dict], maintenance_rows: list[dict]) -> list[dict]:
    rows = [
        {"table_name": row.get("table_name"), "source": "sync", "message": row.get("message")}
        for row in sync_rows
        if row.get("message")
    ]
    rows.extend(
        {
            "table_name": row.get("table_name"),
            "source": "dependency_maintenance",
            "message": row.get("error_message") or row.get("message"),
        }
        for row in maintenance_rows
        if row.get("error_message") or row.get("message")
    )
    return rows


def _format_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    fills = {
        "SUCCESS": PatternFill("solid", fgColor="C6EFCE"),
        "MATCH": PatternFill("solid", fgColor="C6EFCE"),
        "FAILED": PatternFill("solid", fgColor="FFC7CE"),
        "MISMATCH": PatternFill("solid", fgColor="FFC7CE"),
        "WARNING": PatternFill("solid", fgColor="FFEB9C"),
        "SKIPPED": PatternFill("solid", fgColor="FFEB9C"),
    }
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            fill = fills.get(str(cell.value or "").upper())
            if fill:
                cell.fill = fill
    for column in worksheet.columns:
        letter = column[0].column_letter
        width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        worksheet.column_dimensions[letter].width = width


def _flatten_config(value: dict[str, Any], prefix: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for key, item in value.items():
        path = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(item, dict):
            rows.extend(_flatten_config(item, path))
        else:
            rows.append({"key": path, "value": item})
    return rows


def _dataframe(rows: list[dict]) -> pd.DataFrame:
    return pd.DataFrame([_excel_safe_row(row) for row in rows])


def _excel_safe_row(row: dict) -> dict:
    return {key: _excel_safe_value(value) for key, value in row.items()}


def _excel_safe_value(value: Any) -> Any:
    if not isinstance(value, str) or len(value) <= EXCEL_CELL_MAX_CHARS:
        return value
    omitted = len(value) - EXCEL_SAFE_CELL_CHARS
    suffix = f"\n[truncated {omitted} chars for Excel cell limit]"
    return value[: EXCEL_CELL_MAX_CHARS - len(suffix)] + suffix
